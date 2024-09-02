import openpyxl as px
import utilities as ut

psa = [
    "FCFS",
    "SJF",
    "SRTF",
    "Priority (P)",
    "Priority (NP)",
    "Round Robin",
]

excel_file = "data.xlsx"
workbook, worksheet = ut.createExcel(excel_file)


# Fill cells with color in range
def background(color):
    global workbook, worksheet
    for row in worksheet.iter_rows(max_row=100, max_col=100):
        for cell in row:
            cell.fill = px.styles.PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
    workbook.save(excel_file)


def fill_table_structure(worksheet, num_processes, choice, tie_breaker, quantum=None):
    background("f9f991")  # #f9f991

    titles = [
        "Process ID",
        "Arrival Time",
        "Burst Time",
        "Priority",
        "Waiting Time",
        "Response Time",
    ]

    worksheet.title = psa[choice - 1]

    row_spacing, col_spacing = 3, 2  # Set the initial row and column spacing
    fill_titles(titles, row_spacing, col_spacing)

    if choice == 6:  # If the choice is Round Robin, add quantum to the table
        quantum_format(worksheet, quantum, row_spacing=2, col_spacing=2)

    process_cells(
        worksheet, num_processes, titles, row_spacing, col_spacing
    )  # Cell formatting for the processes inputed by the user

    # Last row of the table with the average waiting and response time
    last_row = num_processes + row_spacing + 1
    worksheet.cell(row=last_row, column=col_spacing).value = "Averages"
    for i in range(0, 6):
        cell = worksheet.cell(row=last_row, column=i + col_spacing)

        cell.fill = px.styles.PatternFill(
            start_color="d0d0d0", fill_type="solid"
        )  # Bacground color #d0d0d0
        cell.border = px.styles.Border(
            left=px.styles.Side(style="thin"),
            right=px.styles.Side(style="thin"),
            top=px.styles.Side(style="thin"),
            bottom=px.styles.Side(style="thin"),
        )  # Set the border of the cell to thin

        if i == 0:
            cell.fill = px.styles.PatternFill(
                start_color="000000", fill_type="solid"
            )  # Background color #000000
            cell.font = px.styles.Font(
                color="FFFFFF", bold=True
            )  # Set the font color to white and bold
            cell.alignment = px.styles.Alignment(
                horizontal="center", vertical="center"
            )  # Center the text in the cell

    worksheet.protection.sheet = True
    workbook.save(excel_file)


def process_cells(worksheet, num_processes, titles, row_spacing, col_spacing):
    for row in range(row_spacing + 1, num_processes + row_spacing + 1):
        for col in range(col_spacing, len(titles) + col_spacing):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = px.styles.PatternFill(
                start_color="d0d0d0", fill_type="solid"
            )  # Set the background color of the cell to light grey
            cell.border = px.styles.Border(
                left=px.styles.Side(style="thin"),
                right=px.styles.Side(style="thin"),
                top=px.styles.Side(style="thin"),
                bottom=px.styles.Side(style="thin"),
            )  # Set the border of the cell to thin
            cell.protection = px.styles.Protection(
                locked=False
            )  # Unlock the cell for editing


def quantum_format(worksheet, quantum, row_spacing, col_spacing):
    quantum_title = worksheet.cell(row=row_spacing, column=col_spacing)
    quantum_title.value = "Quantum"
    quantum_title.alignment = px.styles.Alignment(
        horizontal="center", vertical="center"
    )
    quantum_title.fill = px.styles.PatternFill(start_color="000000", fill_type="solid")
    quantum_title.font = px.styles.Font(color="FFFFFF", bold=True)
    quantum_title.border = px.styles.Border(
        left=px.styles.Side(style="thin"),
        right=px.styles.Side(style="thin"),
        top=px.styles.Side(style="thin"),
        bottom=px.styles.Side(style="thin"),
    )

    quantum_cell = worksheet.cell(row=row_spacing, column=col_spacing + 1)
    quantum_cell.value = quantum
    quantum_cell.alignment = px.styles.Alignment(horizontal="center", vertical="center")
    quantum_cell.fill = px.styles.PatternFill(start_color="d0d0d0", fill_type="solid")
    quantum_cell.border = px.styles.Border(
        left=px.styles.Side(style="thin"),
        right=px.styles.Side(style="thin"),
        top=px.styles.Side(style="thin"),
        bottom=px.styles.Side(style="thin"),
    )
    quantum_cell.font = px.styles.Font(color="000000", bold=False)


def fill_titles(titles, row_spacing=1, col_spacing=1):
    # Fill the first row with titles
    for i, title in enumerate(titles):
        cell = worksheet.cell(
            row=row_spacing, column=i + col_spacing
        )  # Assign the cell to a variable
        cell.value = title  # Write the title in the cell
        worksheet.column_dimensions[
            px.utils.get_column_letter(i + col_spacing)
        ].width = (
            len(title) + 5
        )  # Set the width of the column to fit the title text
        cell.alignment = px.styles.Alignment(
            horizontal="center", vertical="center"
        )  # Center the text in the cell
        cell.fill = px.styles.PatternFill(
            start_color="000000", fill_type="solid"
        )  # Set the background color of the cell to black
        cell.font = px.styles.Font(
            color="FFFFFF", bold=True
        )  # Set the font color to white and bold
        cell.border = px.styles.Border(
            left=px.styles.Side(style="thin"),
            right=px.styles.Side(style="thin"),
            top=px.styles.Side(style="thin"),
            bottom=px.styles.Side(style="thin"),
        )
    return titles


