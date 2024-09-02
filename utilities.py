import os
import openpyxl as px
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import subprocess
import sys

def createExcel(excel_file="data.xlsx"):
    if os.path.exists(excel_file):
        deleteExcelFile(excel_file)

    return createExcelFile(excel_file)


def createExcelFile(excel_file="data.xlsx"):
    wb = px.Workbook()
    ws = wb.active
    wb.save(excel_file)
    return wb, ws


def deleteExcelFile(excel_file="data.xlsx"):
    os.remove(excel_file)


# Return a dataframe from the excel file
def loadData(excel_file="data.xlsx"):
    stopExcelIfFileOpen(excel_file)
    df = pd.read_excel(
        excel_file, engine="openpyxl", header=2, index_col=0, usecols=range(1, 7)
    )
    # Drop averages row
    df.drop("Averages", inplace=True, errors="ignore")
    return df


def stopExcelIfFileOpen(excel_file="data.xlsx"):
    if fileCurrentlyOpen(excel_file):
        stopExcel()


# Default row and column spacing is 3; default excel file is "data.xlsx"
def dataframeToExcel(df, row_spacing=3, col_spacing=3, excel_file="data.xlsx"):
    stopExcelIfFileOpen()

    df = getDataframeAverages(df)
    wb, ws = setupEData(excel_file)
    column_count = 0

    for table_column in df.columns.tolist():
        column = df[table_column].tolist()
        column.insert(0, table_column)
        for i in range(len(column)):
            cell = ws.cell(
                row=i + row_spacing,
                column=col_spacing + df.columns.tolist().index(table_column),
            )
            cell.value = column[i] if i != np.nan else ""
        column_count += 1

    indexes = df.index.tolist()  # Get the index names
    indexes.insert(0, "Process ID")  # Insert the index names to the table

    for i in range(len(indexes)):  # Fill the table with the index names
        ws.cell(row=i + row_spacing, column=col_spacing - 1, value=indexes[i])

    wb.save(excel_file)


def setupEData(excel_file="data.xlsx"):
    stopExcelIfFileOpen(excel_file)

    wb = px.load_workbook(excel_file)
    ws = wb.active
    return wb, ws


def stopExcel():
    os.system(
        "taskkill /f /im excel.exe > nul 2>&1"
    )  # close the file if it is open without output on terminal


def openExcel(excel_file="data.xlsx"):
    if sys.platform == "win32":
        os.startfile(excel_file)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, excel_file])


def fileCurrentlyOpen(excel_file="data.xlsx"):
    try:
        wb = px.load_workbook(excel_file)
        if wb is not None:
            return True
        return False
    except px.utils.exceptions.InvalidFileException:
        return True


# Gantt chart generation.
# The variable gantt_chart is a list of lists, where each list contains the start time, Completition Time, and process name.
def ganttChart(gantt_chart, title="Gantt Chart"):
    plt.figure(figsize=(10, 5))
    plt.title(title)
    plt.xlabel("Time")
    plt.ylabel("Process")
    plt.grid(True)
    plt.tight_layout()

    for i, (start, end, name) in enumerate(gantt_chart):
        plt.barh(
            name,
            end - start,
            left=start,
            height=0.5,
            align="center",
            color="skyblue",
            edgecolor="black",
        )  # Create the bar
        plt.text(
            (start + end) / 2,
            name,
            f"{name}",
            ha="center",
            va="center",
            color="black",
            fontsize=8,
        )  # Add the process name to the middle of the bar

    plt.show()


# Calculate the averages for the waiting time, response time, arrival time, and burst time
def getDataframeAverages(dataframe):
    # Insert row if it does not exist
    if "Averages" not in dataframe.index:
        dataframe.loc["Averages"] = np.nan

    for column in dataframe.columns:
        if column != "Priority":
            dataframe.loc["Averages", column] = dataframe[column].mean()
    return dataframe


# Adds the waiting time to the processes that are not running and updates the gantt chart if necessary
def updateWaitingProcesses(
    dataframe, currentProcessID, gantt_chart, time, wait_list, process_list
):
    for p in wait_list:
        if p["Process ID"] != currentProcessID:
            p["Waiting Time"] += 1
            p["Was Executed"] = True

            # If the process was running and is now waiting, update the gantt chart
            if p["Burst Time"] != dataframe.loc[p["Process ID"], "Burst Time"]:
                # If process with the same start time and process ID is in the gantt chart,
                # continue to the next process
                if any(
                    x[0] == p["Start Time"] and x[2] == p["Process ID"]
                    for x in gantt_chart
                ):
                    continue
                else:
                    # Add the process to the gantt chart
                    gantt_chart.append([p["Start Time"], time - 1, p["Process ID"]])

        for i in range(len(process_list)):
            if process_list[i]["Process ID"] == p["Process ID"]:
                process_list[i] = p

    return gantt_chart, process_list
